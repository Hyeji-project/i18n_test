# ===============================
# 1. 라이브러리
# ===============================
from playwright.sync_api import sync_playwright
import pandas as pd
from openpyxl import load_workbook
import re


# ===============================
# 2. 문자열 정규화 (FRD / API 공통)
# ===============================
def normalize_text(text: str) -> str:
    if pd.isna(text) or not text:
        return ""

    text = text.lower()

    # <br>, <br/>, <br /> 제거
    text = re.sub(r"<br\s*/?>", " ", text)
    # &nbsp; 같은 HTML 공백
    text = text.replace("&nbsp;", " ")

    

    # 공백 정리
    text = re.sub(r"\s+", " ", text)

    return text.strip()


# ===============================
# 3. 비교 로직
# ===============================
def compare_i18n(frd: str, api: str) -> str:
    if pd.isna(frd) or pd.isna(api):
        return "SKIP"

    frd = normalize_text(frd)
    api = normalize_text(api)

    # 1️⃣ FRD escape
    pattern = re.escape(frd)

    # 2️⃣ {MM-DD-YYYY}, {0}, {COUNT} → 전부 허용
    pattern = re.sub(r"\\\{.*?\\\}", r".*?", pattern)

    # 3️⃣ 콜론 앞뒤 공백 유연 처리
    pattern = pattern.replace(r"\ :", r"\s*:\s*")

    # 4️⃣ * 제거 (있어도 없어도)
    pattern = pattern.replace(r"\*", r"")

    # 5️⃣ 공백은 전부 느슨하게
    pattern = pattern.replace(r"\ ", r"\s*")

    # 6️⃣ 앞뒤 공백 허용
    pattern = r"\s*" + pattern + r"\s*"

    return "PASS" if re.fullmatch(pattern, api) else "FAIL"


# ===============================
# 4. 엑셀 추출 (모든 시트 / H & K)
# ===============================
def extract_i18n_pairs_all_sheets(excel_path: str):
    wb = load_workbook(excel_path, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        last_frd_id = ""

        for row in ws.iter_rows(min_row=2):
            frd_cell = row[6]    # G열 (FRD ID)
            h_cell = row[7]    # H열 (HQ suggestion_FRD)
            k_cell = row[10]   # K열 (To be filled by local_FRD)

            # 🔹 병합 셀 대응
            if frd_cell.value:
                last_frd_id = str(frd_cell.value).strip()

            frd_id = last_frd_id
            hq_value = h_cell.value
            local_value = k_cell.value

            # H, K 둘 다 값이 있는 경우만
            if hq_value is None or local_value is None:
                continue

            if str(hq_value).strip() == "" or str(local_value).strip() == "":
                continue

            results.append({
                "sheet": sheet_name,
                "frd_id": str(frd_id).strip() if frd_id else "",
                "HQ suggestion_FRD": str(hq_value).strip(),
                "To be filled by Local_FRD": str(local_value).strip()
            })

    return pd.DataFrame(results)


# ===============================
# 5. 메인 로직
# ===============================
def main():
    print("i18n 테스트 시작")

    # 5-1. 엑셀 추출
    df_pairs = extract_i18n_pairs_all_sheets("i18n.xlsx")
    print("엑셀 추출 건수:", len(df_pairs))

     # original(HQ) 기준으로 정보 보존
    excel_dict = {
        row["HQ suggestion_FRD"]: {
            "expected": row["To be filled by Local_FRD"],
            "sheet": row["sheet"],
            "frd_id": row["frd_id"]
        }
        for _, row in df_pairs.iterrows()
    }

    # ===============================
    # 6. Playwright 실행
    # ===============================
    i18n_dict = {}

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        page = browser.new_page()

        def handle_response(response):
             nonlocal i18n_dict 
             url = response.url 

             if "i18n" in url : # 이미 저장했으면 다시 파싱하지 않음 
                if i18n_dict: 
                    return 
                try: 
                    data = response.json() 
                    if isinstance(data, dict): 
                        i18n_dict = data 
                        print("i18n API 인터셉트 성공 :", url) 
                        print("i18n API key 개수:", len(i18n_dict)) 
                except Exception as e: 
                    print("i18n 파싱 실패:", e) 
        page.on("response", handle_response)

        # 메인 페이지 이동
        page.goto("http://hshopfront.samsung.com/vn/")
        # hshopfront 로그인
        page.click('input#username')
        page.fill('input#username', 'qauser')
        page.click('input#password')
        page.fill('input#password', 'qauser1!')
        page.click('button#submit-button')

        #5 테스트 할 페이지 접속
        
        page.wait_for_timeout(1000)#1초 대기
        #쿠키 동의 팝업 클릭 (있을 때만 클릭하도록 수정)
        if page.locator("#truste-consent-button").is_visible(timeout=5000):
            page.click("#truste-consent-button")
        else:
            print("쿠키 동의 팝업이 나타나지 않아 건너뜁니다.")

        #human icon 마우스 오버 (데스크탑/모바일 공용 셀렉터 사용)
        login_btn = page.locator('a.loginBtn:visible, button[an-la="login"]:visible').first
        login_btn.hover()
        page.wait_for_timeout(1000) # 메뉴가 나타날 때까지 짧게 대기
        
        # Sign in 클릭 (더 포괄적인 셀렉터 사용)
        sign_in_link = page.locator('a.loginBtn:visible, a.nv00-gnb-v4__utility-menu--sign-in:visible').last
        sign_in_link.click()
    
        # 이메일 입력
        page.fill("#account", "csrevamp_vn1@teml.net")
        page.locator('button[data-log-id="next"]').click()
        page.fill("#password", "csrevamp1!")
        page.locator('button[data-log-id="signin"]').click()
        page.wait_for_timeout(3000)
        # "Not now" 버튼 처리
        page.click('button[data-testid="test-button-notnow"]') # 첫 번째 버튼은 필수 클릭
        
        # 두 번째 "Not now" 버튼은 있을 때만 클릭 (선택 사항)
        not_now_btn = page.locator('button[data-testid="test-button-notnow"]')
        if not_now_btn.is_visible(timeout=2000):
            not_now_btn.click()
            print("두 번째 'Not now' 버튼 클릭 완료")
        
        page.wait_for_timeout(3000)#3초 대기

        page.goto("https://hshopfront.samsung.com/vn/mypage/")
        page.wait_for_timeout(5000)#초 대기
        browser.close()

    print("최종 i18n API key 개수:", len(i18n_dict))

    # ===============================
    # 7. 엑셀 vs API 비교
    # ===============================
    results = []

    for key, api_value in i18n_dict.items():
        api_value = api_value.strip()

        if key in excel_dict:
            info = excel_dict[key]

            expected = info["expected"].strip()
            sheet = info["sheet"]
            frd_id = info["frd_id"]

            result = compare_i18n(expected, api_value)

            results.append({
                "sheet": sheet,
                "frd_id": frd_id,
                "HQ suggestion_FRD": key,
                "To be filled by Local_FRD": expected,
                "key": key,
                "value": api_value,
                "result": result
            })
        else:
            results.append({
                "sheet": "",
                "frd_id": "",
                "HQ suggestion_FRD": "",
                "To be filled by Local_FRD": "",
                "key": key,
                "value": api_value,
                "result": "Only API"
            })

    # 엑셀에만 존재
    for hq_value, info in excel_dict.items():
        if hq_value not in i18n_dict:
            results.append({
                "sheet": info["sheet"],
                "frd_id": info["frd_id"],
                "HQ suggestion_FRD": hq_value,
                "To be filled by Local_FRD": info["expected"],
                "key": "",
                "value": "",
                "result": "Only FRD"
            })

    # ===============================
    # 8. 결과 저장
    # ===============================
    result_df = pd.DataFrame(
        results,
        columns=[
            "sheet",
            "frd_id",
            "HQ suggestion_FRD",
            "To be filled by Local_FRD",
            "key",
            "value",
            "result"
        ]
    )
    result_df.to_excel("i18n_compare_result.xlsx", index=False)
    print("i18n 비교 결과 엑셀 생성 완료")


# ===============================
# 실행
# ===============================
if __name__ == "__main__":
    main()
